"""
SubAudit — reports/excel_builder.py
Версия спецификации: 2.9
Development Order: Step 5 (Section 16)

generate_excel() → bytes
Используется openpyxl==3.1.2 (Section 15).

Правила экспорта (Section 2):
  - FREE:    Excel недоступен.
  - STARTER: Excel с формулами, без водяного знака.
  - PRO:     Excel с формулами + лист Simulation.

Обязательная проверка плана (Section 2, ⚠):
  Plan MUST be re-verified from Gumroad BEFORE generating any Excel file.
  Вызывающий код (5_dashboard.py, Checkpoint 3) обязан передать актуальный user_plan.

Forecast gate (Section 10):
  forecast_dict = None → лист Forecast не добавляется.
  Если data_months_used < 6 → footer-сообщение в листе Forecast.
"""

from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Цветовая палитра SubAudit (единая для PDF и Excel)
# ---------------------------------------------------------------------------
COLOR_HEADER_BG = "1F3864"   # тёмно-синий заголовок
COLOR_HEADER_FG = "FFFFFF"   # белый текст заголовка
COLOR_SUBHEADER  = "D6E4F0"  # светло-голубой подзаголовок
COLOR_ROW_ALT    = "F2F7FB"  # чередующаяся строка
COLOR_ACCENT     = "2E75B6"  # акцент (синий)
COLOR_GREEN      = "70AD47"  # позитивные метрики
COLOR_RED        = "FF0000"  # предупреждения
COLOR_BORDER     = "BFBFBF"  # цвет рамки ячейки
COLOR_FOOTER_BG  = "EBF3FB"  # фон footer-строки

# ---------------------------------------------------------------------------
# Вспомогательные стили openpyxl
# ---------------------------------------------------------------------------

def _header_font(size: int = 11, bold: bool = True) -> Font:
    return Font(name="Calibri", size=size, bold=bold, color=COLOR_HEADER_FG)


def _body_font(size: int = 10, bold: bool = False, color: str = "000000") -> Font:
    return Font(name="Calibri", size=size, bold=bold, color=color)


def _header_fill() -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=COLOR_HEADER_BG)


def _subheader_fill() -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=COLOR_SUBHEADER)


def _alt_fill() -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=COLOR_ROW_ALT)


def _footer_fill() -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=COLOR_FOOTER_BG)


def _thin_border() -> Border:
    side = Side(style="thin", color=COLOR_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def _center_align() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left_align() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _right_align() -> Alignment:
    return Alignment(horizontal="right", vertical="center")


# ---------------------------------------------------------------------------
# Утилиты форматирования значений
# ---------------------------------------------------------------------------

def _fmt_currency(value: float | None, currency: str = "USD") -> str:
    """Форматирует число как валюту с символом (Section 6 — MRR, ARR и т.д.)."""
    if value is None:
        return "N/A"
    symbols = {"USD": "$", "EUR": "€", "GBP": "£", "RUB": "₽"}
    sym = symbols.get(currency.upper(), currency + " ")
    return f"{sym}{value:,.2f}"


def _fmt_pct(value: float | None) -> str:
    """Форматирует процент (Section 6 — Growth Rate, Churn Rate и т.д.)."""
    if value is None:
        return "N/A"
    return f"{value:.2f}%"


def _fmt_int(value: int | None) -> str:
    if value is None:
        return "N/A"
    return f"{value:,}"


# ---------------------------------------------------------------------------
# Вспомогательные функции стилизации ячеек
# ---------------------------------------------------------------------------

def _style_header_row(ws, row: int, n_cols: int) -> None:
    """Применяет стиль шапки таблицы к строке row."""
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _header_font()
        cell.fill = _header_fill()
        cell.border = _thin_border()
        cell.alignment = _center_align()


def _style_data_row(ws, row: int, n_cols: int, alt: bool = False) -> None:
    """Применяет стиль данных (обычный или чередующийся)."""
    fill = _alt_fill() if alt else PatternFill()
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _body_font()
        if alt:
            cell.fill = fill
        cell.border = _thin_border()
        cell.alignment = _left_align()


def _autofit_columns(ws, min_width: int = 12, max_width: int = 40) -> None:
    """Автоматически подбирает ширину столбцов по содержимому."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(
            min_width, min(max_len + 4, max_width)
        )


# ---------------------------------------------------------------------------
# Sheet 1: Summary — обложка с основными метриками
# ---------------------------------------------------------------------------

def _build_sheet_summary(
    wb: Workbook,
    metrics: dict[str, Any],
    currency: str,
    company_name: dict[str, str],
    user_plan: str,
) -> None:
    """
    Строит лист «Summary» с ключевыми метриками всех блоков.
    Section 9 — Block 1–4 метрики.
    Section 6 — формулы и правила отображения.
    """
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    # --- Заголовок отчёта ---
    display_name = company_name.get("display_name", "SubAudit Report")
    ws.merge_cells("A1:D1")
    title_cell = ws.cell(row=1, column=1, value=f"SubAudit — {display_name}")
    title_cell.font = Font(name="Calibri", size=16, bold=True, color=COLOR_HEADER_FG)
    title_cell.fill = _header_fill()
    title_cell.alignment = _center_align()
    ws.row_dimensions[1].height = 30

    # Дата генерации отчёта
    ws.merge_cells("A2:D2")
    date_str = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    date_cell = ws.cell(row=2, column=1, value=f"Generated: {date_str}  |  Plan: {user_plan.upper()}")
    date_cell.font = _body_font(size=9, color="555555")
    date_cell.alignment = _center_align()
    ws.row_dimensions[2].height = 18

    # --- Заголовки колонок ---
    headers = ["Metric", "Value", "Formula / Note", "Status"]
    current_row = 4
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=h)
    _style_header_row(ws, current_row, len(headers))
    ws.row_dimensions[current_row].height = 20
    current_row += 1

    # ---------------------------------------------------------------------------
    # Block 1 — Revenue (Section 9, Section 6)
    # ---------------------------------------------------------------------------
    block_rows: list[tuple[str, Any, str, str]] = [
        # (label, value, formula_note, status)
        ("── BLOCK 1: REVENUE ──", "", "", ""),
        (
            "MRR",
            _fmt_currency(metrics.get("mrr"), currency),
            "Σ amount (active rows, last_month); multi-row → sum per customer first",
            "✓",
        ),
        (
            "ARR",
            _fmt_currency(metrics.get("arr"), currency),
            "MRR × 12",
            "✓",
        ),
        (
            "ARPU",
            _fmt_currency(metrics.get("arpu"), currency),
            "MRR ÷ unique active customers (last_month). 0 if count==0",
            "✓",
        ),
        (
            "Total Revenue",
            _fmt_currency(metrics.get("total_revenue"), currency),
            "Σ all positive amounts across all time and all statuses",
            "✓",
        ),
        ("── BLOCK 2: GROWTH ──", "", "", ""),
        (
            "New MRR",
            _fmt_currency(metrics.get("new_mrr"), currency),
            "Active rows in last_month WHERE customer is 'new' (first-ever appearance)",
            "✓",
        ),
        (
            "Reactivation MRR",
            _fmt_currency(metrics.get("reactivation_mrr"), currency),
            "Active rows WHERE customer absent 2–9mo then reappears. NOT in New MRR.",
            "✓",
        ),
        (
            "Growth Rate",
            _fmt_pct(metrics.get("growth_rate")),
            "((MRR_last − MRR_prev) ÷ MRR_prev) × 100. N/A if prev==0 or gap.",
            "✓" if metrics.get("growth_rate") is not None else "N/A",
        ),
        (
            "New Subscribers",
            _fmt_int(metrics.get("new_subscribers")),
            "Count of unique new customer_ids (first appearance ever) in last_month",
            "✓",
        ),
        (
            "Reactivated Subscribers",
            _fmt_int(metrics.get("reactivated_subscribers")),
            "Count absent 2–9mo then reappearing. Excluded if max(amount) > 6× ARPU.",
            "✓",
        ),
        ("── BLOCK 3: RETENTION ──", "", "", ""),
        (
            "Churn Rate",
            _fmt_pct(metrics.get("churn_rate")),
            "(lost_subs ÷ active_prev_month) × 100. N/A if prev==0 or gap.",
            "✓" if metrics.get("churn_rate") is not None else "N/A",
        ),
        (
            "Revenue Churn",
            _fmt_currency(metrics.get("revenue_churn"), currency),
            "Scenarios A/B/C/D — see spec Section 8. Refund not double-counted.",
            "✓",
        ),
        (
            "NRR",
            _fmt_pct(metrics.get("nrr")),
            "CLAMP(((MRR_last − rev_churn + exp_mrr) ÷ MRR_prev) × 100, 0, 999)",
            "⚠ >200%" if (metrics.get("nrr") or 0) > 200 else "✓",
        ),
        ("── BLOCK 4: HEALTH ──", "", "", ""),
        (
            "LTV",
            _fmt_currency(metrics.get("ltv"), currency),
            "ARPU ÷ (churn_rate / 100). Cap = ARPU×36 if churn==0. NOT for unit economics.",
            "✓",
        ),
        (
            "Active Subscribers",
            _fmt_int(metrics.get("active_subscribers")),
            "Unique customer_ids with status=='active' in last_month",
            "✓",
        ),
        (
            "Lost Subscribers",
            _fmt_int(metrics.get("lost_subscribers")),
            "Active in prev_month, absent from active rows in last_month",
            "✓" if metrics.get("lost_subscribers") is not None else "N/A",
        ),
        (
            "Existing Subscribers",
            _fmt_int(metrics.get("existing_subscribers")),
            "Active in BOTH prev_month AND last_month",
            "✓",
        ),
    ]

    for idx, (label, value, note, status) in enumerate(block_rows):
        is_block_label = label.startswith("──")
        ws.cell(row=current_row, column=1, value=label)
        ws.cell(row=current_row, column=2, value=value)
        ws.cell(row=current_row, column=3, value=note)
        ws.cell(row=current_row, column=4, value=status)

        if is_block_label:
            # Строка-разделитель блока
            ws.merge_cells(
                start_row=current_row, start_column=1,
                end_row=current_row, end_column=4,
            )
            cell = ws.cell(row=current_row, column=1)
            cell.font = Font(name="Calibri", size=10, bold=True, color=COLOR_ACCENT)
            cell.fill = _subheader_fill()
            cell.alignment = _center_align()
        else:
            alt = (idx % 2 == 0)
            _style_data_row(ws, current_row, 4, alt=alt)
            # Выделяем предупреждения красным
            if status.startswith("⚠"):
                ws.cell(row=current_row, column=4).font = Font(
                    name="Calibri", size=10, bold=True, color=COLOR_RED
                )

        ws.row_dimensions[current_row].height = 18
        current_row += 1

    # NRR > 200% предупреждение (Section 6 — ⚠ NRR tooltip)
    nrr_val = metrics.get("nrr")
    if nrr_val is not None and nrr_val > 200:
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row, end_column=4,
        )
        warn_cell = ws.cell(
            row=current_row, column=1,
            value=(
                "⚠ NRR exceeds 200% — likely caused by limited prior-month data. "
                "Interpret with caution."
            ),
        )
        warn_cell.font = Font(name="Calibri", size=9, italic=True, color=COLOR_RED)
        warn_cell.alignment = _left_align()
        current_row += 1

    # LTV cap предупреждение (Section 6 — LTV cap tooltip)
    current_row += 1
    ws.merge_cells(
        start_row=current_row, start_column=1,
        end_row=current_row, end_column=4,
    )
    ltv_note = ws.cell(
        row=current_row, column=1,
        value=(
            "ℹ LTV cap = 36 months (ARPU × 36). "
            "Do NOT use capped LTV for unit economics or CAC payback decisions."
        ),
    )
    ltv_note.font = Font(name="Calibri", size=9, italic=True, color="555555")
    ltv_note.alignment = _left_align()

    # Ширина столбцов
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 12
    ws.freeze_panes = "A5"


# ---------------------------------------------------------------------------
# Sheet 2: Metrics Detail — полный список с Excel-формулами
# ---------------------------------------------------------------------------

def _build_sheet_metrics_detail(
    wb: Workbook,
    metrics: dict[str, Any],
    currency: str,
) -> None:
    """
    Лист «Metrics Detail» — числовые значения + Excel-формулы
    для самостоятельной верификации пользователем.
    Section 2: Starter/PRO — 'with formulas'.
    Section 6: формулы метрик как документация.
    """
    ws = wb.create_sheet(title="Metrics Detail")
    ws.sheet_view.showGridLines = False

    headers = ["Metric Key", "Raw Value", "Formatted", "Excel Formula (verification)"]
    ws.cell(row=1, column=1, value="Metrics — Raw Values & Verification Formulas")
    ws.cell(row=1, column=1).font = Font(name="Calibri", size=13, bold=True, color=COLOR_ACCENT)
    ws.merge_cells("A1:D1")

    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=2, column=col_idx, value=h)
    _style_header_row(ws, 2, len(headers))

    # Метрики с числовыми значениями и Excel-формулами
    # Формулы ссылаются на ячейки колонки B (Raw Value), демонстрируя логику из Section 6
    # Строки: MRR=3, ARR=4, ARPU=5, Total Revenue=6, New MRR=7, Reactivation MRR=8,
    #         Growth Rate=9, New Subscribers=10, Reactivated Subscribers=11,
    #         Churn Rate=12, Revenue Churn=13, NRR=14, LTV=15, Active Subscribers=16,
    #         Lost Subscribers=17, Existing Subscribers=18,
    #         [Промежуточные] MRR Prev Month=19, Active Subs Prev Month=20, Expansion MRR=21
    rows = [
        ("mrr",                  metrics.get("mrr"),                  _fmt_currency(metrics.get("mrr"), currency),        "Calculated from raw data"),
        ("arr",                  metrics.get("arr"),                  _fmt_currency(metrics.get("arr"), currency),        "=B3*12"),
        ("arpu",                 metrics.get("arpu"),                 _fmt_currency(metrics.get("arpu"), currency),       "=IF(B16=0,0,B3/B16)"),
        ("total_revenue",        metrics.get("total_revenue"),        _fmt_currency(metrics.get("total_revenue"), currency), "Calculated from raw data"),
        ("new_mrr",              metrics.get("new_mrr"),              _fmt_currency(metrics.get("new_mrr"), currency),    "Calculated from raw data"),
        ("reactivation_mrr",     metrics.get("reactivation_mrr"),     _fmt_currency(metrics.get("reactivation_mrr"), currency), "Calculated from raw data"),
        ("growth_rate",          metrics.get("growth_rate"),          _fmt_pct(metrics.get("growth_rate")),               "=IF(B19=0,\"N/A\",(B3-B19)/B19*100)"),
        ("new_subscribers",      metrics.get("new_subscribers"),      _fmt_int(metrics.get("new_subscribers")),           "Calculated from raw data"),
        ("reactivated_subscribers", metrics.get("reactivated_subscribers"), _fmt_int(metrics.get("reactivated_subscribers")), "Calculated from raw data"),
        ("churn_rate",           metrics.get("churn_rate"),           _fmt_pct(metrics.get("churn_rate")),                "=IF(B20=0,\"N/A\",B17/B20*100)"),
        ("revenue_churn",        metrics.get("revenue_churn"),        _fmt_currency(metrics.get("revenue_churn"), currency), "Calculated from raw data"),
        ("nrr",                  metrics.get("nrr"),                  _fmt_pct(metrics.get("nrr")),                       "=MAX(0,MIN(999,(B3-B13+B21)/B19*100))"),
        ("ltv",                  metrics.get("ltv"),                  _fmt_currency(metrics.get("ltv"), currency),        "=IF(B12=0,B5*36,B5/(B12/100))"),
        ("active_subscribers",   metrics.get("active_subscribers"),   _fmt_int(metrics.get("active_subscribers")),        "Calculated from raw data"),
        ("lost_subscribers",     metrics.get("lost_subscribers"),     _fmt_int(metrics.get("lost_subscribers")),          "=B20-B18"),
        ("existing_subscribers", metrics.get("existing_subscribers"), _fmt_int(metrics.get("existing_subscribers")),      "Calculated from raw data"),
        # Промежуточные значения для формул (не отображаются в Summary)
        ("mrr_prev_month",       metrics.get("mrr_prev_month"),       _fmt_currency(metrics.get("mrr_prev_month"), currency), "MRR of previous month"),
        ("active_subscribers_prev_month", metrics.get("active_subscribers_prev_month"), _fmt_int(metrics.get("active_subscribers_prev_month")), "Active subscribers in previous month"),
        ("expansion_mrr",        metrics.get("expansion_mrr"),        _fmt_currency(metrics.get("expansion_mrr"), currency), "MRR increase from existing customers"),
    ]

    for idx, (key, raw_val, fmt_val, formula) in enumerate(rows):
        r = idx + 3
        ws.cell(row=r, column=1, value=key)
        # Числовое значение — None → пустая строка, иначе float/int
        ws.cell(row=r, column=2, value=raw_val if raw_val is not None else "N/A")
        ws.cell(row=r, column=3, value=fmt_val)
        ws.cell(row=r, column=4, value=formula)
        _style_data_row(ws, r, 4, alt=(idx % 2 == 0))

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 55
    ws.freeze_panes = "A3"


# ---------------------------------------------------------------------------
# Sheet 3: Cohort Table (Section 7)
# ---------------------------------------------------------------------------

def _build_sheet_cohort(
    wb: Workbook,
    cohort_df: pd.DataFrame | None,
) -> None:
    """
    Лист «Cohort Table».
    Section 7: max 12 когорт, retention% = retained_N / cohort_size × 100.
    Цветовая раскраска RdYlGn (красный→жёлтый→зелёный).
    Нет данных / None → лист с информационным сообщением.
    """
    ws = wb.create_sheet(title="Cohort Table")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    title_cell = ws.cell(row=1, column=1, value="Cohort Retention Analysis")
    title_cell.font = Font(name="Calibri", size=13, bold=True, color=COLOR_ACCENT)

    if cohort_df is None or cohort_df.empty:
        # Section 7: 3 distinct cohort months required → else return None
        ws.cell(row=3, column=1,
                value="ℹ Cohort table requires at least 3 distinct cohort months of data.")
        ws.cell(row=3, column=1).font = Font(name="Calibri", size=10, italic=True, color="555555")
        return

    # Section 7: Max 12 most recent cohorts
    df_display = cohort_df.tail(12)

    # Заголовок: Cohort + Month 0, Month 1, ...
    cols = list(df_display.columns)  # первая колонка — cohort label, остальные — Month N
    headers_row = 3
    for col_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=headers_row, column=col_idx, value=str(col_name))
    _style_header_row(ws, headers_row, len(cols))

    # Данные когорт + условное форматирование цветом (RdYlGn)
    for r_idx, (_, row_data) in enumerate(df_display.iterrows()):
        r = headers_row + 1 + r_idx
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r, column=c_idx)
            if isinstance(val, float) and not pd.isna(val):
                # Section 7: Retention % = retained_N / cohort_size × 100 → значения 0–100.
                # openpyxl number_format "0.00%" умножает хранимое значение на 100 при отображении.
                # Поэтому сохраняем val / 100 (десятичная дробь), чтобы Excel показал корректный %:
                # 75.5 → 0.755 → отображается как 75.50%  (без деления было бы 7550.00% — БАГ)
                cell.value = round(val / 100.0, 4)
                cell.number_format = "0.00%"
                # RdYlGn: красный ≤ 20%, жёлтый ≈ 50%, зелёный ≥ 80%
                pct = val / 100.0 if val > 1 else val  # нормализуем для цвета
                cell.fill = _rdylgn_fill(pct)
            else:
                cell.value = val
                cell.fill = _alt_fill() if r_idx % 2 == 0 else PatternFill()
            cell.font = _body_font()
            cell.border = _thin_border()
            cell.alignment = _center_align()

    _autofit_columns(ws)
    ws.freeze_panes = "B4"


def _rdylgn_fill(pct: float) -> PatternFill:
    """
    Аппроксимация палитры RdYlGn для openpyxl.
    pct: 0.0 → красный, 0.5 → жёлтый, 1.0 → зелёный.
    Section 7: background_gradient(colormap='RdYlGn').
    """
    pct = max(0.0, min(1.0, pct))
    if pct <= 0.5:
        # Красный → Жёлтый
        t = pct * 2
        r = 255
        g = int(255 * t)
        b = 0
    else:
        # Жёлтый → Зелёный
        t = (pct - 0.5) * 2
        r = int(255 * (1 - t))
        g = 200
        b = 0
    hex_color = f"{r:02X}{g:02X}{b:02X}"
    return PatternFill(fill_type="solid", fgColor=hex_color)


# ---------------------------------------------------------------------------
# Sheet 4: Forecast (Section 10)
# ---------------------------------------------------------------------------

def _build_sheet_forecast(
    wb: Workbook,
    forecast_dict: dict[str, Any] | None,
    currency: str,
) -> None:
    """
    Лист «Forecast».
    Section 10 forecast gate:
      - forecast_dict is None → футер "Forecast omitted — fewer than 6 months of data available."
      - data_months_used < 6 → только Realistic, без экспорта; показываем предупреждение.
      - ≥ 6 months → все 3 сценария.
    Section 2: гейт плана уже проверен вызывающим кодом.
    """
    ws = wb.create_sheet(title="Forecast")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    title_cell = ws.cell(row=1, column=1, value="Revenue Forecast — 12 Months Forward")
    title_cell.font = Font(name="Calibri", size=13, bold=True, color=COLOR_ACCENT)

    # --- Forecast gate (Section 10) ---
    if forecast_dict is None:
        ws.merge_cells("A3:E3")
        footer_cell = ws.cell(
            row=3, column=1,
            value="Forecast omitted — fewer than 6 months of data available.",
        )
        footer_cell.font = Font(name="Calibri", size=10, italic=True, color="555555")
        footer_cell.fill = _footer_fill()
        footer_cell.alignment = _left_align()
        return

    data_months_used: int = forecast_dict.get("data_months_used", 0)
    has_full_scenarios: bool = data_months_used >= 6  # Section 10: ≥ 6 months → all 3 scenarios

    # Предупреждение для 3–5 месяцев (Section 10: Show st.warning() BEFORE chart)
    current_row = 3
    if not has_full_scenarios:
        ws.merge_cells(f"A{current_row}:E{current_row}")
        warn_cell = ws.cell(
            row=current_row, column=1,
            value=(
                f"⚠ Forecast based on {data_months_used} months of data "
                "(3–5 months). Realistic scenario only. "
                "Statistical reliability is limited — interpret with caution."
            ),
        )
        warn_cell.font = Font(name="Calibri", size=9, italic=True, color=COLOR_RED)
        warn_cell.fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
        warn_cell.alignment = _left_align()
        current_row += 2

    # --- Заголовки таблицы ---
    if has_full_scenarios:
        headers = ["Month", "Period", "Pessimistic", "Realistic", "Optimistic"]
    else:
        headers = ["Month", "Period", "Realistic"]

    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=current_row, column=col_idx, value=h)
    _style_header_row(ws, current_row, len(headers))
    current_row += 1

    # --- Данные прогноза ---
    # forecast_dict ожидается с ключами: periods, realistic, pessimistic (опц.), optimistic (опц.)
    periods: list[str] = forecast_dict.get("periods", [])
    realistic: list[float] = forecast_dict.get("realistic", [])
    pessimistic: list[float] = forecast_dict.get("pessimistic", []) if has_full_scenarios else []
    optimistic: list[float] = forecast_dict.get("optimistic", []) if has_full_scenarios else []

    symbols = {"USD": "$", "EUR": "€", "GBP": "£", "RUB": "₽"}
    curr_sym = symbols.get(currency.upper(), currency + " ")

    for idx, period in enumerate(periods):
        real_val = realistic[idx] if idx < len(realistic) else None
        pess_val = pessimistic[idx] if idx < len(pessimistic) else None
        opti_val = optimistic[idx] if idx < len(optimistic) else None

        ws.cell(row=current_row, column=1, value=idx + 1)
        ws.cell(row=current_row, column=2, value=period)

        if has_full_scenarios:
            ws.cell(row=current_row, column=3,
                    value=round(pess_val, 2) if pess_val is not None else "N/A")
            ws.cell(row=current_row, column=4,
                    value=round(real_val, 2) if real_val is not None else "N/A")
            ws.cell(row=current_row, column=5,
                    value=round(opti_val, 2) if opti_val is not None else "N/A")
            for col in range(3, 6):
                cell = ws.cell(row=current_row, column=col)
                cell.number_format = f'"{curr_sym}"#,##0.00'
        else:
            ws.cell(row=current_row, column=3,
                    value=round(real_val, 2) if real_val is not None else "N/A")
            cell = ws.cell(row=current_row, column=3)
            cell.number_format = f'"{curr_sym}"#,##0.00'

        _style_data_row(ws, current_row, len(headers), alt=(idx % 2 == 0))
        current_row += 1

    # --- Footer: Negative guard note (Section 10) ---
    current_row += 1
    ws.merge_cells(f"A{current_row}:E{current_row}")
    footer_cell = ws.cell(
        row=current_row, column=1,
        value=(
            "ℹ All forecast values are clamped to ≥ 0. "
            "Pessimistic = projected × (1 − churn_rate/100 × 1.20). "
            "Optimistic = projected × (1 − churn_rate/100 × 0.80). "
            "Churn fallback = 5% if churn_rate unavailable."
        ),
    )
    footer_cell.font = Font(name="Calibri", size=9, italic=True, color="555555")
    footer_cell.fill = _footer_fill()
    footer_cell.alignment = _left_align()

    _autofit_columns(ws)
    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# Sheet 5: Simulation — PRO only (Section 11)
# ---------------------------------------------------------------------------

def _build_sheet_simulation(
    wb: Workbook,
    simulation_dict: dict[str, Any] | None,
    currency: str,
) -> None:
    """
    Лист «Simulation» — только для плана PRO (Section 2, Section 11).
    Отображает входные параметры, MRR-прогноз по месяцам и предупреждение
    об однородности ARPU (KNOWN LIMITATION, Section 11).
    """
    ws = wb.create_sheet(title="Simulation (PRO)")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    title_cell = ws.cell(row=1, column=1, value="Subscription Simulator — PRO Feature")
    title_cell.font = Font(name="Calibri", size=13, bold=True, color=COLOR_ACCENT)

    if simulation_dict is None:
        ws.cell(row=3, column=1,
                value="ℹ Simulation data is not available for this session.")
        ws.cell(row=3, column=1).font = Font(
            name="Calibri", size=10, italic=True, color="555555"
        )
        return

    # --- Входные параметры (Section 11 — Input Parameters) ---
    params_start = 3
    ws.cell(row=params_start, column=1, value="Input Parameters").font = Font(
        name="Calibri", size=11, bold=True, color=COLOR_HEADER_FG
    )
    ws.cell(row=params_start, column=1).fill = _header_fill()
    ws.merge_cells(f"A{params_start}:B{params_start}")

    param_rows = [
        ("Churn Reduction",          simulation_dict.get("churn_reduction"),     "0.0–1.0 (0.20 = 20% less churn)"),
        ("New Customers / Month",    simulation_dict.get("new_customers_month"),  "0–10,000"),
        ("Price Increase (ARPU Δ)",  simulation_dict.get("price_increase"),      "-0.5–5.0 (0.10 = +10%)"),
        ("Base MRR",                 simulation_dict.get("base_mrr"),            f"Starting MRR ({currency})"),
        ("Base ARPU",                simulation_dict.get("base_arpu"),           f"Starting ARPU ({currency})"),
    ]

    for idx, (label, val, note) in enumerate(param_rows):
        r = params_start + 1 + idx
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=val if val is not None else "N/A")
        ws.cell(row=r, column=3, value=note)
        _style_data_row(ws, r, 3, alt=(idx % 2 == 0))

    # --- MRR-прогноз по месяцам ---
    sim_data_start = params_start + len(param_rows) + 3
    ws.cell(row=sim_data_start, column=1, value="Projected MRR by Month").font = Font(
        name="Calibri", size=11, bold=True, color=COLOR_HEADER_FG
    )
    ws.cell(row=sim_data_start, column=1).fill = _header_fill()
    ws.merge_cells(f"A{sim_data_start}:G{sim_data_start}")

    sim_headers = ["Month", "MRR", "MRR Change", "MRR Change %", "Active Subscribers",
                   "New Customers Added", "Churn Rate (effective)"]
    for col_idx, h in enumerate(sim_headers, start=1):
        ws.cell(row=sim_data_start + 1, column=col_idx, value=h)
    _style_header_row(ws, sim_data_start + 1, len(sim_headers))

    monthly_data: list[dict] = simulation_dict.get("monthly_data", [])
    symbols_map = {"USD": "$", "EUR": "€", "GBP": "£", "RUB": "₽"}
    curr_sym = symbols_map.get(currency.upper(), currency + " ")

    for idx, month_row in enumerate(monthly_data):
        r = sim_data_start + 2 + idx
        mrr_val = month_row.get("mrr")
        mrr_change = month_row.get("mrr_change")
        mrr_change_pct = month_row.get("mrr_change_pct")  # None если base_mrr==0 (Section 11)

        ws.cell(row=r, column=1, value=idx + 1)
        ws.cell(row=r, column=2, value=round(mrr_val, 2) if mrr_val is not None else "N/A")
        ws.cell(row=r, column=3, value=round(mrr_change, 2) if mrr_change is not None else "N/A")
        # Section 11: mrr_change_pct = None when base_mrr == 0 → display "N/A"
        ws.cell(row=r, column=4,
                value=f"{mrr_change_pct:.2f}%" if mrr_change_pct is not None else "N/A")
        ws.cell(row=r, column=5, value=month_row.get("active_subscribers", "N/A"))
        ws.cell(row=r, column=6, value=month_row.get("new_customers_added", "N/A"))
        ws.cell(row=r, column=7,
                value=_fmt_pct(month_row.get("effective_churn_rate")))

        # Форматирование валюты
        for col in (2, 3):
            ws.cell(row=r, column=col).number_format = f'"{curr_sym}"#,##0.00'

        _style_data_row(ws, r, len(sim_headers), alt=(idx % 2 == 0))

    # Excel SUM-формула для итогового MRR (Section 2: "with formulas")
    if monthly_data:
        sum_row = sim_data_start + 2 + len(monthly_data)
        ws.cell(row=sum_row, column=1, value="TOTAL / AVERAGE")
        data_col_start = sim_data_start + 2
        data_col_end = sum_row - 1
        ws.cell(row=sum_row, column=2,
                value=f"=AVERAGE(B{data_col_start}:B{data_col_end})")
        ws.cell(row=sum_row, column=3,
                value=f"=SUM(C{data_col_start}:C{data_col_end})")
        ws.cell(row=sum_row, column=2).number_format = f'"{curr_sym}"#,##0.00'
        ws.cell(row=sum_row, column=3).number_format = f'"{curr_sym}"#,##0.00'
        for col in range(1, len(sim_headers) + 1):
            cell = ws.cell(row=sum_row, column=col)
            cell.font = Font(name="Calibri", size=10, bold=True)
            cell.fill = PatternFill(fill_type="solid", fgColor=COLOR_SUBHEADER)
            cell.border = _thin_border()

    # --- ARPU homogeneity warning (Section 11 — KNOWN LIMITATION) ---
    footer_row = sim_data_start + 2 + len(monthly_data) + 2
    ws.merge_cells(f"A{footer_row}:G{footer_row}")
    arpu_warn = ws.cell(
        row=footer_row, column=1,
        value=(
            "⚠ Results assume uniform ARPU. With mixed pricing tiers, "
            "actual revenue impact may differ by 30–60%. "
            "Mixed-tier modelling is v2 roadmap."
        ),
    )
    arpu_warn.font = Font(name="Calibri", size=9, italic=True, color=COLOR_RED)
    arpu_warn.fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    arpu_warn.alignment = _left_align()

    _autofit_columns(ws)
    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# Sheet 6: Data Quality Flags (Section 14 — data_quality_flags)
# ---------------------------------------------------------------------------

def _build_sheet_data_quality(
    wb: Workbook,
    data_quality_flags: dict[str, Any] | None,
) -> None:
    """
    Лист «Data Quality» — информация о prev_month_status,
    last_month_is_fallback, last_month_used.
    Section 14: data_quality_flags = {prev_month_status, last_month_is_fallback, last_month_used}.
    """
    ws = wb.create_sheet(title="Data Quality")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:C1")
    ws.cell(row=1, column=1, value="Data Quality Flags").font = Font(
        name="Calibri", size=13, bold=True, color=COLOR_ACCENT
    )

    if not data_quality_flags:
        ws.cell(row=3, column=1, value="No data quality flags available.")
        return

    flag_rows = [
        ("prev_month_status",
         data_quality_flags.get("prev_month_status", "unknown"),
         "'ok' = Previous month data available. 'gap' = Missing calendar month. "
         "'missing' = No data before last month. "
         "Affects: Growth Rate, Churn Rate, NRR, Lost Subscribers (shown as N/A when not 'ok')."),
        ("last_month_is_fallback",
         str(data_quality_flags.get("last_month_is_fallback", False)),
         "True = Last month has fewer than 5 active customers (data quality warning shown). "
         "False = Last month meets minimum threshold for reliable metrics."),
        ("last_month_used",
         str(data_quality_flags.get("last_month_used", "N/A")),
         "The calendar month (YYYY-MM) used as 'last month' for all metric calculations. "
         "This is the most recent month with sufficient data."),
    ]

    headers = ["Flag", "Value", "Meaning"]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=3, column=col_idx, value=h)
    _style_header_row(ws, 3, len(headers))

    for idx, (flag, value, meaning) in enumerate(flag_rows):
        r = 4 + idx
        ws.cell(row=r, column=1, value=flag)
        ws.cell(row=r, column=2, value=value)
        ws.cell(row=r, column=3, value=meaning)
        _style_data_row(ws, r, 3, alt=(idx % 2 == 0))
        ws.row_dimensions[r].height = 22

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 65


# ---------------------------------------------------------------------------
# Основная функция: generate_excel() → bytes
# ---------------------------------------------------------------------------

def generate_excel(
    metrics_dict: dict[str, Any],
    forecast_dict: dict[str, Any] | None,
    simulation_dict: dict[str, Any] | None,
    cohort_df: pd.DataFrame | None,
    data_quality_flags: dict[str, Any] | None,
    user_plan: str,
    currency: str,
    company_name: dict[str, str],
) -> bytes:
    """
    Генерирует Excel-отчёт и возвращает его как bytes.

    Параметры (из session_state согласно Section 14):
      metrics_dict       — dict из get_all_metrics(df_clean)
      forecast_dict      — dict или None; None → forecast gate (Section 10)
      simulation_dict    — dict или None; только для PRO (Section 2, 11)
      cohort_df          — DataFrame или None (Section 7)
      data_quality_flags — dict (Section 14)
      user_plan          — 'free' / 'starter' / 'pro' (Section 2)
      currency           — str, например 'USD'
      company_name       — {'display_name': str, 'filename_safe_name': str} (Section 14)

    Возвращает bytes — содержимое .xlsx файла.

    ⚠ ВАЖНО (Section 2):
      Plan MUST be re-verified from Gumroad BEFORE вызовом этой функции.
      Вызывающий код (5_dashboard.py, Checkpoint 3) отвечает за проверку.

    ⚠ FREE (Section 2):
      Excel export недоступен для FREE плана.
      Вызывающий код должен заблокировать доступ до вызова этой функции.
    """
    # Защитный guard: FREE план не должен попасть сюда,
    # но на случай некорректного вызова — явная проверка
    if user_plan == "free":
        raise PermissionError(
            "Excel export is not available on the FREE plan. "
            "Upgrade to STARTER or PRO. (Section 2)"
        )

    wb = Workbook()

    # Sheet 1: Summary (Block 1–4 метрики)
    _build_sheet_summary(wb, metrics_dict, currency, company_name, user_plan)

    # Sheet 2: Metrics Detail с Excel-формулами (Starter + PRO, Section 2)
    _build_sheet_metrics_detail(wb, metrics_dict, currency)

    # Sheet 3: Cohort Table (Section 7)
    _build_sheet_cohort(wb, cohort_df)

    # Sheet 4: Forecast с gate (Section 10)
    _build_sheet_forecast(wb, forecast_dict, currency)

    # Sheet 5: Simulation — только PRO (Section 2, Section 11)
    if user_plan == "pro":
        _build_sheet_simulation(wb, simulation_dict, currency)

    # Sheet 6: Data Quality Flags (Section 14)
    _build_sheet_data_quality(wb, data_quality_flags)

    # Сериализация в bytes (openpyxl → BytesIO)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
